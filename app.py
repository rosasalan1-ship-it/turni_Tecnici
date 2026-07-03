import streamlit as st
import pandas as pd
from ortools.sat.python import cp_model
 
# =========================
# CONFIG
# =========================
TECHS = ["Giugno", "De Simone", "Ragusa", "Cunsolo"]   # tecnici interni gestiti dal solver
EXTERNAL = "Esterno"                                    # tecnico esterno: SOLO input manuale, mai assegnato dal solver
ALL_ROWS = TECHS + [EXTERNAL]                           # tutte le righe mostrate nelle tabelle
DAYS_DEFAULT = 30   # usato solo finché non è stata scelta una data di inizio
WORK_SHIFTS = ["M", "P", "N"]
REP_SHIFTS = ["REPN", "REPD"]
ALL_SHIFTS = WORK_SHIFTS + REP_SHIFTS
 
MAX_REP_MONTH = 7          # REP totali (REPN+REPD) al mese, per tutti i tecnici interni
GIUGNO_MAX_NOTTI = 6       # N + REPN per il tecnico nella riga "Giugno" (era Tiziana)
 
# simbolo colorato per ogni riga: solo VISIVO, aiuta a non perdere la riga durante lo
# scroll orizzontale sui giorni. Inserito nelle celle vuote della griglia di INPUT e
# sempre rimosso prima che il dato venga salvato/letto dal solver (vedi normalizza_cella).
ROW_DOT = {
    "Giugno": "🟦",
    "De Simone": "🟩",
    "Ragusa": "🟨",
    "Cunsolo": "🟧",
    EXTERNAL: "⬜",
}
 
 
def normalizza_cella(val):
    """Rimuove il puntino visivo (se presente) e gli spazi: usato OVUNQUE il dato
    venga letto/salvato, cosi' il puntino non interferisce mai con la logica."""
    val = str(val)
    for dot in ROW_DOT.values():
        val = val.replace(dot, "")
    return val.strip()
 
 
def aggiungi_puntini_per_display(df):
    """Ritorna una COPIA del dataframe con il puntino colorato della riga inserito
    nelle celle vuote, solo per la visualizzazione nella griglia di input."""
    out = df.copy()
    for tecnico in out.index:
        dot = ROW_DOT.get(tecnico, "")
        if not dot:
            continue
        for col in out.columns:
            if normalizza_cella(out.loc[tecnico, col]) == "":
                out.loc[tecnico, col] = dot
    return out
 
 
st.set_page_config(page_title="Gestione Turni", layout="wide")
 
# =========================
# AUTENTICAZIONE
# =========================
PASSWORD = "Radiologia_Tecnici2026"
 
if "autenticato" not in st.session_state:
    st.session_state.autenticato = False
 
if not st.session_state.autenticato:
    st.title("🔒 Accesso riservato")
    pwd = st.text_input("Password:", type="password", placeholder="Inserisci la password")
    if st.button("Accedi"):
        if pwd == PASSWORD:
            st.session_state.autenticato = True
            st.rerun()
        else:
            st.error("❌ Password errata.")
    st.stop()  # blocca tutto il resto dell'app finché non si è autenticati
 
st.title("📅 Pianificazione Turni - Tecnici")
 
# CSS globale: bordi su tutte le tabelle/data_editor dell'app
st.markdown(
    """
    <style>
    div[data-testid="stDataFrame"] table,
    div[data-testid="stDataFrame"] th,
    div[data-testid="stDataFrame"] td,
    div[data-testid="stDataEditor"] table,
    div[data-testid="stDataEditor"] th,
    div[data-testid="stDataEditor"] td {
        border: 1px solid #444 !important;
        border-collapse: collapse !important;
    }
    div[data-testid="stDataFrame"] [role="grid"],
    div[data-testid="stDataEditor"] [role="grid"] {
        border: 1px solid #444 !important;
    }
    div[data-testid="stDataFrame"] [role="gridcell"],
    div[data-testid="stDataFrame"] [role="columnheader"],
    div[data-testid="stDataEditor"] [role="gridcell"],
    div[data-testid="stDataEditor"] [role="columnheader"] {
        border: 1px solid #444 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)
 
st.markdown(
    """
**Codici disponibili per ogni giorno:**
`F` · `MAL` · `RC` · `CS` · `RR` · `IM` · `IP` · `IN` · `IREPN` · `IREPD` · `DM` · `DP` · `DN` · `DREPN` · `DREPD` · `M` · `P` · `N` · `REPN` · `REPD`
 
ℹ️ Riga **Esterno**: solo `M`, `P`, `N`, `REPN`, `REPD` (assegnazione manuale, mai automatica). Cella vuota = quel giorno è come se non esistesse.
"""
)
 
# =========================
# INPUT - TABELLA EDITABILE
# =========================
st.subheader("Inserisci disponibilità / vincoli")
 
data_inizio = st.date_input("Data di inizio del periodo (giorno 1)", value=None)
 
GIORNI_SETT = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]
 
# numero di giorni nel mese del 1° giorno scelto (28/29/30/31). Se nessuna data
# è stata scelta, si usa un valore di default finché l'utente non la imposta.
if data_inizio is not None:
    import calendar
    DAYS = calendar.monthrange(data_inizio.year, data_inizio.month)[1]
else:
    DAYS = DAYS_DEFAULT
 
COL_NAMES = [str(i) for i in range(1, DAYS + 1)]
 
if "inp" not in st.session_state:
    st.session_state.inp = pd.DataFrame("", index=ALL_ROWS, columns=COL_NAMES)
elif st.session_state.inp.shape[1] != DAYS:
    old = st.session_state.inp
    new = pd.DataFrame("", index=ALL_ROWS, columns=COL_NAMES)
    common_cols = [c for c in old.columns if c in new.columns]
    new[common_cols] = old[common_cols]
    st.session_state.inp = new
    st.warning(
        f"⚠️ Il mese selezionato ha {DAYS} giorni: la griglia è stata adattata "
        f"(i dati inseriti nei giorni comuni sono stati mantenuti)."
    )
 
if st.session_state.get("_appena_salvato"):
    st.success("✅ Tabella salvata.")
    st.session_state._appena_salvato = False
 
if "festivo_giorni" not in st.session_state:
    st.session_state.festivo_giorni = ""
 
# festivi già SALVATI in precedenza (usati per marcare le intestazioni della griglia di input)
_festivi_set_saved = set()
for _tok in st.session_state.festivo_giorni.replace(" ", "").split(","):
    if _tok.isdigit():
        _festivi_set_saved.add(int(_tok))
 
if data_inizio is not None:
    import datetime
    col_labels = []
    for i in range(DAYS):
        giorno = data_inizio + datetime.timedelta(days=i)
        is_dom_o_festivo = giorno.weekday() == 6 or (i + 1) in _festivi_set_saved
        prefix = "🔴" if is_dom_o_festivo else ""
        col_labels.append(f"{prefix}{i+1}\n{GIORNI_SETT[giorno.weekday()]}")
else:
    st.info("Seleziona la data di inizio per vedere i giorni della settimana nella tabella.")
    col_labels = COL_NAMES
 
# Bottone Pulisci: cancella anche la cache interna del widget (non solo il dataframe)
col_clear, _ = st.columns([1, 5])
with col_clear:
    if st.button("🧹 Pulisci griglia"):
        st.session_state.inp = pd.DataFrame("", index=ALL_ROWS, columns=COL_NAMES)
        st.session_state.festivo_giorni = ""
        editor_key = f"editor_input_{DAYS}"
        if editor_key in st.session_state:
            del st.session_state[editor_key]
        st.rerun()
 
# La griglia è dentro un FORM: le modifiche si fissano solo al click di "Salva",
# evitando che Streamlit "perda" un'edit perché ricarica la pagina troppo presto
# mentre stai ancora scrivendo nelle celle.
with st.form("form_input", border=False):
    festivo_input = st.text_input(
        "Giorni festivi (numeri separati da virgola) — si comportano come domenica:",
        value=st.session_state.festivo_giorni,
        placeholder="es: 5, 12, 25",
    )
 
    display_df = aggiungi_puntini_per_display(st.session_state.inp)
    display_df.columns = col_labels
 
    # colonne strette (max 4 lettere nei codici): si vede tutto il mese senza scroll
    col_config_input = {
        col: st.column_config.TextColumn(label=col, width=46)
        for col in display_df.columns
    }
 
    edited_display = st.data_editor(
        display_df,
        use_container_width=False,
        num_rows="fixed",
        key=f"editor_input_{DAYS}",
        column_config=col_config_input,
    )
 
    salva = st.form_submit_button("💾 Salva tabella", type="secondary")
 
if salva:
    edited = edited_display.copy()
    edited.columns = COL_NAMES
    for _col in edited.columns:
        edited[_col] = edited[_col].apply(normalizza_cella)
    st.session_state.inp = edited
    st.session_state.festivo_giorni = festivo_input
    st.session_state._appena_salvato = True
    st.rerun()  # forza ricaricamento immediato dal dato salvato, evita perdita celle
 
 
inp = st.session_state.inp
 
# parsing dei giorni festivi inseriti come testo (es. "5,12,25")
_festivi_set = set()
for _tok in st.session_state.festivo_giorni.replace(" ", "").split(","):
    if _tok.isdigit():
        _festivi_set.add(int(_tok))
 
st.divider()
 
 
# =========================
# BOTTONE SOLVE
# =========================
if st.button("🚀 Genera turni", type="primary"):
 
    model = cp_model.CpModel()
 
    # giorno della settimana basato sulla data di inizio scelta dall'utente
    import datetime as _dt
 
    def weekday_of(d):
        if data_inizio is None:
            return d % 7  # fallback: assume giorno1 = lunedi
        return (data_inizio + _dt.timedelta(days=d)).weekday()
 
    def is_saturday(d):
        return weekday_of(d) == 5
 
    def is_sunday(d):
        return weekday_of(d) == 6
 
    def is_friday(d):
        return weekday_of(d) == 4
 
    def is_festivo(d):
        return (d + 1) in _festivi_set
 
    def is_giorno_speciale(d):
        # domenica O giorno segnato FESTIVO: si comporta come domenica (solo REPD/REPN)
        return is_sunday(d) or is_festivo(d)
 
    # =========================
    # VARIABILI
    # =========================
    x = {}
    for t in TECHS:
        for d in range(DAYS):
            for s in ALL_SHIFTS:
                if is_giorno_speciale(d):
                    if s not in ("REPN", "REPD"):
                        continue  # domenica/festivo: solo REPD (diurna) e REPN (notturna), niente M/P/N
                else:
                    if s == "REPD":
                        continue  # REPD esiste solo la domenica/festivo
                x[(t, d, s)] = model.NewBoolVar(f"x_{t}_{d}_{s}")
 
    def var(t, d, s):
        return x.get((t, d, s))
 
    # =========================
    # FUNZIONI INPUT
    # =========================
    def is_absent(v):
        v = str(v).upper()
        return any(code in v for code in ["F", "MAL", "RC", "CS", "RR"])
 
    def is_blocked(v, s):
        return f"I{s}" in str(v)
 
    def is_preferred(v, s):
        return f"D{s}" in str(v)
 
    def is_forced(v, s):
        # se la cella contiene ESATTAMENTE il codice del turno (es. "M", "REPN")
        # senza prefisso D/I, è un'assegnazione OBBLIGATA
        return str(v).strip() == s
 
    # =========================
    # 0b. TECNICO ESTERNO (riga "Esterno"): SOLO input manuale, mai assegnato dal solver.
    #     Se per un dato giorno/turno l'Esterno e' assegnato (cella = "M"/"P"/"N"/"REPN"/"REPD"),
    #     quel turno viene bloccato per TUTTI i tecnici interni quel giorno (niente doppia copertura).
    #     Se la cella dell'Esterno e' vuota quel giorno, e' come se non esistesse: i tecnici interni
    #     restano liberi di coprire normalmente quel turno.
    # =========================
    external_assigned = {}  # (d, s) -> True se l'Esterno copre quel turno quel giorno
    for d in range(DAYS):
        cella_esterno = str(inp.loc[EXTERNAL, str(d + 1)]) if EXTERNAL in inp.index else ""
        for s in ALL_SHIFTS:
            if is_forced(cella_esterno, s):
                external_assigned[(d, s)] = True
                for t in TECHS:
                    v_ = var(t, d, s)
                    if v_ is not None:
                        model.Add(v_ == 0)
 
    # =========================
    # 1. MUTUAL EXCLUSION GIORNALIERA
    # =========================
    for t in TECHS:
        for d in range(DAYS):
            mpn = [var(t, d, s) for s in ["M", "P", "N"] if var(t, d, s) is not None]
            if mpn:
                model.Add(sum(mpn) <= 1)
 
            repn = var(t, d, "REPN")
            n_ = var(t, d, "N")
            if repn is not None and n_ is not None:
                model.Add(repn + n_ <= 1)
 
            repd = var(t, d, "REPD")
            if repd is not None:
                for s in ["M", "P", "N"]:
                    v_ = var(t, d, s)
                    if v_ is not None:
                        model.Add(repd + v_ <= 1)
 
    # =========================
    # 2. FERIE / MALATTIA -> HARD BLOCK su tutto
    # =========================
    for t in TECHS:
        for d in range(DAYS):
            if is_absent(inp.loc[t, str(d + 1)]):
                for s in ALL_SHIFTS:
                    v_ = var(t, d, s)
                    if v_ is not None:
                        model.Add(v_ == 0)
 
    # =========================
    # 3. INDISPONIBILITÀ TURNO (I*) -> HARD BLOCK
    # =========================
    for t in TECHS:
        for d in range(DAYS):
            v = inp.loc[t, str(d + 1)]
            for s in ALL_SHIFTS:
                if is_blocked(v, s):
                    v_ = var(t, d, s)
                    if v_ is not None:
                        model.Add(v_ == 0)
 
    # =========================
    # 3b. ASSEGNAZIONE FORZATA (cella = codice turno puro, es. "M", "REPN")
    # =========================
    for t in TECHS:
        for d in range(DAYS):
            v = inp.loc[t, str(d + 1)]
            for s in ALL_SHIFTS:
                if is_forced(v, s):
                    v_ = var(t, d, s)
                    if v_ is not None:
                        model.Add(v_ == 1)
 
    # =========================
    # 4. UNITA' LAVORATIVE SETTIMANALI (M=1, P=1, N=2 unita'; REP non conta).
    #    Per ogni tecnico, in ogni settimana (piena o parziale, es. gli ultimi giorni
    #    del mese), le unita' lavorate devono essere ESATTAMENTE pari ai giorni
    #    lavorativi della settimana per quel tecnico, dove:
    #      giorni_lavorativi = giorni della settimana
    #                          - domeniche/festivi (non si contano mai)
    #                          - giorni di assenza F/MAL/RC/CS/RR del tecnico (si sottraggono)
    #    Esempio: settimana di 7 giorni con 1 domenica -> 6 giorni lavorativi -> 6 unita'.
    #    Esempio: settimana parziale di 3 giorni (nessuna domenica/assenza) -> 3 unita'.
    # =========================
    weekly_unit_vars = []  # somma_unita_var per ogni (tecnico, settimana): usata anche nell'obiettivo
    for t in TECHS:
        for w_start in range(0, DAYS, 7):
            days_in_week = range(w_start, min(w_start + 7, DAYS))
            terms = []
            giorni_lavorativi = 0
            for d in days_in_week:
                if is_giorno_speciale(d):
                    continue  # domenica/festivo: non e' un giorno lavorativo, non conta
                cella = inp.loc[t, str(d + 1)]
                if is_absent(cella):
                    continue  # F/MAL/RC/CS/RR: giorno sottratto dal conteggio per questo tecnico
                giorni_lavorativi += 1
                if var(t, d, "M") is not None:
                    terms.append(var(t, d, "M"))
                if var(t, d, "P") is not None:
                    terms.append(var(t, d, "P"))
                n_ = var(t, d, "N")
                if n_ is not None:
                    terms.append(2 * n_)
            if terms or giorni_lavorativi:
                model.Add(sum(terms) == giorni_lavorativi)
                weekly_unit_vars.append(sum(terms))
 
    # =========================
    # 5. MAX 7 REP AL MESE (REPN + REPD) PER TUTTI
    # =========================
    rep_total = {}
    for t in TECHS:
        terms = []
        for d in range(DAYS):
            for s in REP_SHIFTS:
                v_ = var(t, d, s)
                if v_ is not None:
                    terms.append(v_)
        rep_total[t] = sum(terms)
        model.Add(rep_total[t] <= MAX_REP_MONTH)
 
    # =========================
    # 6. GIUGNO: MAX 6 NOTTI (N + REPN) AL MESE (era "Tiziana", stessa limitazione per riga)
    # =========================
    giugno_notti = []
    for d in range(DAYS):
        n_ = var("Giugno", d, "N")
        if n_ is not None:
            giugno_notti.append(n_)
        repn_ = var("Giugno", d, "REPN")
        if repn_ is not None:
            giugno_notti.append(repn_)
    model.Add(sum(giugno_notti) <= GIUGNO_MAX_NOTTI)
 
    # =========================
    # 7. RIPOSO DOPO N: STOP COMPLETO IL GIORNO DOPO (nessun turno/REP)
    #    Il rientro avviene il giorno successivo ancora, in REP o M/P
    # =========================
    for t in TECHS:
        for d in range(DAYS - 1):
            n_ = var(t, d, "N")
            if n_ is None:
                continue
            for s in ALL_SHIFTS:
                v_next = var(t, d + 1, s)
                if v_next is not None:
                    model.Add(v_next == 0).OnlyEnforceIf(n_)
 
    # =========================
    # 7b. DOPO LA REP NOTTURNA NON SI FA LA MATTINA, BENSI' IL POMERIGGIO
    # =========================
    for t in TECHS:
        for d in range(DAYS - 1):
            repn_ = var(t, d, "REPN")
            m_next = var(t, d + 1, "M")
            if repn_ is not None and m_next is not None:
                model.Add(m_next == 0).OnlyEnforceIf(repn_)
 
    # NOTA: la REP di per se' non prevede un giorno di riposo completo
    # (solo dopo la N c'e' lo stop totale); dopo la REPN si esclude solo la M.
 
    # =========================
    # 9. COPERTURA M E P: SEMPRE ALMENO 1 PERSONA (HARD), MAX 2 IN M
    # =========================
    for d in range(DAYS):
        m_vars = [var(t, d, "M") for t in TECHS if var(t, d, "M") is not None]
        p_vars = [var(t, d, "P") for t in TECHS if var(t, d, "P") is not None]
        if m_vars:
            model.Add(sum(m_vars) >= 1)
            # nessun tetto rigido: se avanzano operatori senza altro turno,
            # vengono messi in M (gestito come bonus nell'obiettivo)
        if p_vars:
            model.Add(sum(p_vars) >= 1)
 
    # =========================
    # 10. COPERTURA NOTTE: ESATTAMENTE 1 OPERATORE (REPN o N, mai più di uno)
    #     COPERTURA REPD DOMENICA: ESATTAMENTE 1 OPERATORE
    #     (la copertura minima è SOFT: si accetta scopertura solo se causata
    #      da troppe F/MAL; ma se qualcuno la copre, deve essere uno solo - HARD)
    # =========================
    night_uncovered = {}
    sunday_repd_uncovered = {}
 
    for d in range(DAYS):
        # copertura REPD diurna (domenica o giorno festivo)
        if is_giorno_speciale(d):
            repd_vars = [var(t, d, "REPD") for t in TECHS if var(t, d, "REPD") is not None]
            if repd_vars:
                model.Add(sum(repd_vars) <= 1)  # un solo operatore interno in REPD
                if external_assigned.get((d, "REPD")):
                    # l'Esterno coprirà la REPD: già coperta, nessuno slack di scopertura
                    sunday_repd_uncovered[d] = None
                else:
                    slack_d = model.NewBoolVar(f"repd_uncovered_{d}")
                    model.Add(sum(repd_vars) + slack_d >= 1)
                    sunday_repd_uncovered[d] = slack_d
 
        # copertura notte (REPN o N) - tutti i giorni, incluse le domeniche (solo REPN)
        coverage_vars = []
        for t in TECHS:
            n_ = var(t, d, "N")
            if n_ is not None:
                coverage_vars.append(n_)
            repn_ = var(t, d, "REPN")
            if repn_ is not None:
                coverage_vars.append(repn_)
        if coverage_vars:
            model.Add(sum(coverage_vars) <= 1)  # un solo operatore interno copre la notte
            if external_assigned.get((d, "REPN")) or external_assigned.get((d, "N")):
                # l'Esterno coprirà la notte: già coperta, nessuno slack di scopertura
                night_uncovered[d] = None
            else:
                slack = model.NewBoolVar(f"night_uncovered_{d}")
                model.Add(sum(coverage_vars) + slack >= 1)
                night_uncovered[d] = slack
 
    # =========================
    # 11. CONTINUITA' WEEKEND (SOFT) + REPD A OPERATORE DIVERSO (HARD)
    #     chi fa REPN sabato è preferito per la REPN di domenica (stessa persona)
    #     la REPD di domenica deve invece andare a un altro operatore
    # =========================
    weekend_bonus_vars = []
    for t in TECHS:
        for d in range(DAYS - 1):
            if is_saturday(d):
                sat_repn = var(t, d, "REPN")
                sun_repn = var(t, d + 1, "REPN")
                sun_repd = var(t, d + 1, "REPD")
 
                # bonus soft: stessa persona REPN sabato e REPN domenica
                if sat_repn is not None and sun_repn is not None:
                    both = model.NewBoolVar(f"weekend_cont_{t}_{d}")
                    model.AddMultiplicationEquality(both, [sat_repn, sun_repn])
                    weekend_bonus_vars.append(both)
 
                # hard: chi fa REPN sabato NON può fare la REPD di domenica
                if sat_repn is not None and sun_repd is not None:
                    model.Add(sat_repn + sun_repd <= 1)
 
    # =========================
    # 11c. CHI FA VENERDI' NOTTE FA SABATO POMERIGGIO E DOMENICA REPD
    #      (rotazione: un weekend impegnato, uno libero)
    # =========================
    for t in TECHS:
        for d in range(DAYS - 2):
            if is_friday(d):
                fri_n = var(t, d, "N")
                sat_p = var(t, d + 1, "P")
                sun_repd = var(t, d + 2, "REPD")
                if fri_n is not None and sat_p is not None:
                    model.Add(sat_p == 1).OnlyEnforceIf(fri_n)
                if fri_n is not None and sun_repd is not None:
                    model.Add(sun_repd == 1).OnlyEnforceIf(fri_n)
 
    # =========================
    # 11d. NOTTI/REP NON CONSECUTIVE (SOFT): penalita' se la stessa persona
    #      fa notte (N o REPN) due giorni di seguito
    # =========================
    consecutive_night_vars = []
    for t in TECHS:
        for d in range(DAYS - 1):
            today_terms = [v for v in [var(t, d, "N"), var(t, d, "REPN")] if v is not None]
            tomorrow_terms = [v for v in [var(t, d + 1, "N"), var(t, d + 1, "REPN")] if v is not None]
            if not today_terms or not tomorrow_terms:
                continue
            today_night = model.NewBoolVar(f"night_today_{t}_{d}")
            model.AddMaxEquality(today_night, today_terms)
            tomorrow_night = model.NewBoolVar(f"night_tomorrow_{t}_{d}")
            model.AddMaxEquality(tomorrow_night, tomorrow_terms)
            both = model.NewBoolVar(f"consec_night_{t}_{d}")
            model.AddMultiplicationEquality(both, [today_night, tomorrow_night])
            consecutive_night_vars.append(both)
    # =========================
    # 11e. DOMENICA/FESTIVO: REPN e REPD NON POSSONO ESSERE DELLO STESSO OPERATORE
    # =========================
    for t in TECHS:
        for d in range(DAYS):
            if is_giorno_speciale(d):
                repn_ = var(t, d, "REPN")
                repd_ = var(t, d, "REPD")
                if repn_ is not None and repd_ is not None:
                    model.Add(repn_ + repd_ <= 1)
 
    # =========================
    # 11b. DISTRIBUZIONE EQUA DI REPN, REPD, N TRA GLI OPERATORI
    #      minimizziamo lo squilibrio (max - min) per ciascun tipo di turno
    # =========================
    def fairness_penalty(shift_name, label):
        """Crea variabili max/min sul totale di 'shift_name' per tecnico
        e ritorna una lista di termini obiettivo che penalizzano lo squilibrio."""
        totals = []
        for t in TECHS:
            terms = [var(t, d, shift_name) for d in range(DAYS) if var(t, d, shift_name) is not None]
            if not terms:
                continue
            tot = model.NewIntVar(0, DAYS, f"tot_{label}_{t}")
            model.Add(tot == sum(terms))
            totals.append(tot)
        if len(totals) < 2:
            return []
        max_v = model.NewIntVar(0, DAYS, f"max_{label}")
        min_v = model.NewIntVar(0, DAYS, f"min_{label}")
        model.AddMaxEquality(max_v, totals)
        model.AddMinEquality(min_v, totals)
        spread = model.NewIntVar(0, DAYS, f"spread_{label}")
        model.Add(spread == max_v - min_v)
        return [spread]
 
    fairness_terms = []
    fairness_terms += fairness_penalty("REPN", "repn")
    fairness_terms += fairness_penalty("REPD", "repd")
    fairness_terms += fairness_penalty("N", "n")
 
    # =========================
    # 12. OBIETTIVO
    # =========================
    objective = []
 
    for t in TECHS:
        for d in range(DAYS):
            v = inp.loc[t, str(d + 1)]
            for s in ALL_SHIFTS:
                if is_preferred(v, s):
                    v_ = var(t, d, s)
                    if v_ is not None:
                        objective.append(5 * v_)
 
    for t in TECHS:
        for d in range(DAYS):
            repn_ = var(t, d, "REPN")
            if repn_ is not None:
                objective.append(3 * repn_)
            repd_ = var(t, d, "REPD")
            if repd_ is not None:
                objective.append(3 * repd_)
            n_ = var(t, d, "N")
            if n_ is not None:
                objective.append(1 * n_)
 
    objective += [2 * v for v in weekend_bonus_vars]
 
    # NOTA: le unita' settimanali sono ora fissate da un vincolo di uguaglianza HARD
    # (sezione 4 sopra), quindi non serve piu' un bonus soft nell'obiettivo per spingerle.
 
    # bonus generale di riempimento: qualsiasi M/P assegnato è meglio di un buco
    # (i blank non coperti da F/MAL/I*/D* vengono preferibilmente messi in M)
    for d in range(DAYS):
        m_vars = [var(t, d, "M") for t in TECHS if var(t, d, "M") is not None]
        p_vars = [var(t, d, "P") for t in TECHS if var(t, d, "P") is not None]
        objective += [2 * v for v in m_vars]
        objective += [1 * v for v in p_vars]  # leggera preferenza per M rispetto a P sui "buchi"
 
    # bonus: dopo la REP notturna, preferiamo il pomeriggio (non obbligatorio, ma incentivato)
    for t in TECHS:
        for d in range(DAYS - 1):
            repn_ = var(t, d, "REPN")
            p_next = var(t, d + 1, "P")
            if repn_ is not None and p_next is not None:
                both = model.NewBoolVar(f"repn_then_p_{t}_{d}")
                model.AddMultiplicationEquality(both, [repn_, p_next])
                objective.append(2 * both)
 
    # bonus: lo stesso giorno della REPN, preferiamo che il tecnico faccia anche M mattina
    # (non obbligatorio). var(t, d, "M") e' None nei giorni domenica/festivo (dove M non
    # esiste nel modello), quindi il bonus si applica automaticamente solo dove e' possibile.
    # M e REPN non sono mutuamente esclusivi nel modello, quindi possono coesistere.
    for t in TECHS:
        for d in range(DAYS):
            repn_ = var(t, d, "REPN")
            m_ = var(t, d, "M")
            if repn_ is not None and m_ is not None:
                both = model.NewBoolVar(f"m_with_repn_{t}_{d}")
                model.AddMultiplicationEquality(both, [m_, repn_])
                objective.append(4 * both)  # peso 4: preferenza forte ma non blocca la soluzione
 
    objective += [-50 * s for s in night_uncovered.values() if s is not None]
    objective += [-50 * s for s in sunday_repd_uncovered.values() if s is not None]
 
    # penalita' forte sullo squilibrio tra operatori (equita')
    objective += [-4 * s for s in fairness_terms]
 
    # penalita' per notti/REP consecutive (preferiamo distribuirle nel mese)
    objective += [-3 * s for s in consecutive_night_vars]
 
    model.Maximize(sum(objective))
 
    # =========================
    # SOLVER
    # =========================
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 30
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)
 
    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        out = pd.DataFrame("", index=ALL_ROWS, columns=[str(i) for i in range(1, DAYS + 1)])
        for t in TECHS:
            for d in range(DAYS):
                cella_input = str(inp.loc[t, str(d + 1)])
                if is_absent(cella_input):
                    codice_assenza = next(
                        (c for c in ["F", "MAL", "RC", "CS", "RR"] if c in cella_input.upper()),
                        "F",
                    )
                    out.loc[t, str(d + 1)] = codice_assenza
                    continue
                assigned = []
                for s in ALL_SHIFTS:
                    v_ = var(t, d, s)
                    if v_ is not None and solver.Value(v_) == 1:
                        assigned.append(s)
                out.loc[t, str(d + 1)] = "+".join(assigned)
 
        # riga Esterno: riportata 1:1 da quanto inserito manualmente (mai dal solver)
        for d in range(DAYS):
            cella_esterno = str(inp.loc[EXTERNAL, str(d + 1)]).strip() if EXTERNAL in inp.index else ""
            out.loc[EXTERNAL, str(d + 1)] = cella_esterno if cella_esterno in ALL_SHIFTS else ""
 
        # giorni scoperti -> riga speciale "SCOPERTO" sotto la tabella
        giorni_scoperti_notte = [d + 1 for d, s in night_uncovered.items() if s is not None and solver.Value(s) == 1]
        giorni_scoperti_repd = [d + 1 for d, s in sunday_repd_uncovered.items() if s is not None and solver.Value(s) == 1]
 
        # riga aggiuntiva nella tabella che segnala "SCOPERTO" sul giorno
        out.loc["⚠️ SCOPERTO NOTTE"] = ""
        for gd in giorni_scoperti_notte:
            out.loc["⚠️ SCOPERTO NOTTE", str(gd)] = "SCOPERTO"
        out.loc["⚠️ SCOPERTO REP DOMENICA"] = ""
        for gd in giorni_scoperti_repd:
            out.loc["⚠️ SCOPERTO REP DOMENICA", str(gd)] = "SCOPERTO"
 
        st.success(
            "Soluzione "
            + ("ottimale" if status == cp_model.OPTIMAL else "trovata (non ottimale)")
            + " ✅"
        )
 
        if giorni_scoperti_notte:
            st.warning(f"⚠️ Notti scoperte nei giorni: {', '.join(map(str, giorni_scoperti_notte))}")
        if giorni_scoperti_repd:
            st.warning(f"⚠️ REP diurna domenicale scoperta nei giorni: {', '.join(map(str, giorni_scoperti_repd))}")
 
        BORDER = "border: 1px solid #444;"
 
        def color_shift(val):
            if val == "SCOPERTO":
                return BORDER + "background-color: #f8d7da; color: #842029; font-weight: bold"
            if val in ("F", "MAL"):
                return BORDER + "background-color: #e2e3e5; color: #6c757d; font-style: italic"
            if "REPD" in val:
                return BORDER + "background-color: #d4edda"
            if "REPN" in val:
                return BORDER + "background-color: #ffe5b4"
            if "N" in val:
                return BORDER + "background-color: #d6d8d9"
            if "M" in val:
                return BORDER + "background-color: #cce5ff"
            if "P" in val:
                return BORDER + "background-color: #fff3cd"
            return BORDER
 
        # intestazioni con i giorni della settimana (🔴 davanti al numero = domenica/festivo)
        col_labels_out = []
        for i in range(DAYS):
            prefix = "🔴" if is_giorno_speciale(i) else ""
            if data_inizio is not None:
                giorno = data_inizio + _dt.timedelta(days=i)
                col_labels_out.append(f"{prefix}{i+1}\n{GIORNI_SETT[giorno.weekday()]}")
            else:
                col_labels_out.append(f"{prefix}{i + 1}")
 
        out_display = out.copy()
        out_display.columns = col_labels_out
 
        col_config_out = {
            col: st.column_config.TextColumn(label=col, width=46)
            for col in out_display.columns
        }
 
        st.subheader("Turni generati")
        st.dataframe(
            out_display.style.map(color_shift).set_properties(**{"text-align": "center"}),
            use_container_width=False,
            column_config=col_config_out,
        )
 
        st.subheader("Riepilogo per tecnico")
        ABSENCE_CODES = ["F", "MAL", "RC", "CS", "RR"]
        RIEPILOGO_COLS = ALL_SHIFTS + ABSENCE_CODES
        summary = pd.DataFrame(index=ALL_ROWS, columns=RIEPILOGO_COLS).fillna(0)
        for t in ALL_ROWS:
            for s in ALL_SHIFTS:
                summary.loc[t, s] = sum(
                    1 for d in range(DAYS) if s in str(out.loc[t, str(d + 1)]).split("+")
                )
            for s in ABSENCE_CODES:
                summary.loc[t, s] = sum(
                    1 for d in range(DAYS) if str(out.loc[t, str(d + 1)]) == s
                )
        st.dataframe(
            summary.style.set_properties(**{"border": "1px solid #444", "text-align": "center"}),
            use_container_width=False,
        )
 
        # =========================
        # EXPORT CSV ED EXCEL
        # =========================
        csv = out_display.to_csv().encode("utf-8")
        st.download_button(
            "⬇️ Scarica turni in CSV",
            data=csv,
            file_name="turni_generati.csv",
            mime="text/csv",
        )
 
        import io
        from openpyxl.styles import Border, Side, Alignment, Font
        from openpyxl.worksheet.page import PageMargins
        from openpyxl.utils import get_column_letter
 
        MESI_IT = [
            "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
            "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre",
        ]
        nome_mese = MESI_IT[data_inizio.month - 1] if data_inizio is not None else ""
        intestazione_testo = (
            f"ASP Caltanissetta - P.O. Suor Cecilia Bassarocco - Radiologia - Mese di {nome_mese}"
            if nome_mese
            else "ASP Caltanissetta - P.O. Suor Cecilia Bassarocco - Radiologia"
        )
 
        # tabella trasposta: giorni in riga, tecnici (incluso Esterno) in colonna
        out_vert = out_display.copy().transpose()
 
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            HEADER_ROWS = 2  # 1 riga di intestazione + 1 riga vuota di distacco
            out_display.to_excel(writer, sheet_name="Turni", startrow=HEADER_ROWS)
            summary.to_excel(writer, sheet_name="Riepilogo", startrow=HEADER_ROWS)
            out_vert.to_excel(writer, sheet_name="Turni (verticale)", startrow=HEADER_ROWS)
 
            thin = Side(border_style="thin", color="444444")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            title_font = Font(bold=True, size=12)
            signature_font = Font(size=10)
 
            def applica_intestazione_e_firme(ws, ultima_riga_tabella):
                # intestazione testuale fissa nelle prime righe del foglio (non header di stampa)
                n_col_title = max(2, ws.max_column)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_col_title)
                ws.cell(row=1, column=1, value=intestazione_testo).font = title_font
 
                # spazio firme: 4 tecnici interni + Responsabile, due righe sotto la tabella
                riga = ultima_riga_tabella + 2
                ws.cell(row=riga, column=1, value="Firme:").font = Font(bold=True)
                riga += 1
                n_col_merge = min(6, max(2, ws.max_column))  # estende la linea firma su alcune colonne
                for tecnico in TECHS:
                    ws.cell(row=riga, column=1, value=f"{tecnico}:").font = signature_font
                    ws.merge_cells(start_row=riga, start_column=2, end_row=riga, end_column=n_col_merge)
                    ws.cell(row=riga, column=2, value="_______________________________")
                    riga += 1
                ws.cell(row=riga, column=1, value="Responsabile:").font = Font(bold=True, size=10)
                ws.merge_cells(start_row=riga, start_column=2, end_row=riga, end_column=n_col_merge)
                ws.cell(row=riga, column=2, value="_______________________________")
 
            for sheet_name in ["Turni", "Riepilogo", "Turni (verticale)"]:
                ws = writer.sheets[sheet_name]
                header_row_excel = HEADER_ROWS + 1  # riga in cui inizia l'intestazione della tabella dati
 
                # bordi su tutte le celle della tabella dati (incluse intestazioni)
                for row in ws.iter_rows(
                    min_row=header_row_excel, max_row=ws.max_row, min_col=1, max_col=ws.max_column
                ):
                    for cell in row:
                        cell.border = border
 
                # intestazioni tabella: centra e va a capo
                for cell in ws[header_row_excel]:
                    cell.alignment = header_align
 
                # larghezza colonne e impostazioni di stampa specifiche per foglio
                if sheet_name == "Turni":
                    ws.column_dimensions["A"].width = 12  # colonna nomi tecnici
                    for col_idx in range(2, ws.max_column + 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = 6
                    ws.page_setup.orientation = "landscape"
                    ws.page_setup.fitToWidth = 1
                    ws.page_setup.fitToHeight = 0
                elif sheet_name == "Riepilogo":
                    ws.column_dimensions["A"].width = 14
                    ws.page_setup.orientation = "landscape"
                    ws.page_setup.fitToWidth = 1
                    ws.page_setup.fitToHeight = 0
                else:  # Turni (verticale): tecnici in colonna, giorni in riga -> stampa verticale
                    ws.column_dimensions["A"].width = 14  # colonna giorni
                    for col_idx in range(2, ws.max_column + 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = 16
                    ws.page_setup.orientation = "portrait"
                    ws.page_setup.fitToWidth = 1
                    ws.page_setup.fitToHeight = 0
 
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                ws.page_margins = PageMargins(
                    left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.2, footer=0.2
                )
 
                applica_intestazione_e_firme(ws, ws.max_row)
        buffer.seek(0)
 
        st.download_button(
            "⬇️ Scarica turni in Excel (.xlsx)",
            data=buffer,
            file_name="turni_generati.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
 
    else:
        st.error("❌ Nessuna soluzione trovata. Controlla i vincoli inseriti (troppe ferie/indisponibilità rigide?).")